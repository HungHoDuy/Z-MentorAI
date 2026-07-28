import os
import sys
import time
import json
import logging
import uuid
import datetime
import re
import asyncio
from typing import Any, Optional, List
from pathlib import Path
from io import BytesIO

from fastapi import FastAPI, HTTPException, Response, Request
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv
from sse_starlette.sse import EventSourceResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
env_path = Path(__file__).resolve().parent.parent / ".env"
if env_path.exists():
    load_dotenv(dotenv_path=env_path)
    print(f"Academic Architect: Loaded environment from {env_path}")
else:
    print("Academic Architect: Warning, no .env file found.")

from google.cloud import firestore
from google.cloud.firestore_v1.vector import Vector
from google.cloud.firestore_v1.base_vector_query import DistanceMeasure
from langchain_google_vertexai import VertexAIEmbeddings

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("academic_architect")

COLLECTION_NAME = "learning_material"
JOBS_VECTOR_COLLECTION = "data_vector_embeddings"
GANTT_COLLECTION_NAME = "academic_gantt_charts"

# Initialize LLM
llm = None
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
USE_VERTEX_AI = os.getenv("USE_VERTEX_AI", "false").lower() == "true"
USE_FIRESTORE = os.getenv("USE_FIRESTORE", "false").lower() == "true"

if USE_VERTEX_AI:
    from langchain_google_vertexai import ChatVertexAI
    llm = ChatVertexAI(
        model_name="gemini-3.5-flash",
        location="asia-southeast1",
        temperature=0.7
    )
    logger.info("Academic Architect: Using Vertex AI gemini-3.5-flash")
elif GEMINI_API_KEY:
    from langchain_google_genai import ChatGoogleGenerativeAI
    llm = ChatGoogleGenerativeAI(model="gemini-3.5-flash", temperature=0.7)
    logger.info("Academic Architect: Using ChatGoogleGenerativeAI with API Key")
else:
    logger.warning("Academic Architect: Neither Vertex AI nor GEMINI_API_KEY is configured for LLM.")

app = FastAPI(title="Academic Architect Agent")

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global in-memory SSE event queues for task progress streaming
progress_queues: dict[str, asyncio.Queue] = {}

async def emit_progress(task_id: Optional[str], step: str, status: str, duration_ms: float, message: str):
    """Emit a structured SSE progress event with duration timing."""
    if not task_id:
        return
    if task_id not in progress_queues:
        progress_queues[task_id] = asyncio.Queue()
        
    event_payload = {
        "timestamp": datetime.datetime.utcnow().isoformat() + "Z",
        "step": step,
        "status": status,
        "duration_ms": round(duration_ms, 2),
        "message": message
    }
    await progress_queues[task_id].put(event_payload)
    logger.info(f"Progress [{task_id}] ({step} - {duration_ms:.1f}ms): {message}")

# --- Pydantic Data Schemas ---
class SearchRequest(BaseModel):
    search_text: str
    search_mode: str  # "name" or "description"
    domain: Optional[str] = None
    k: Optional[int] = 3

class SearchResponse(BaseModel):
    feature: str = "course_search"
    search_text: str
    search_mode: str
    domain: Optional[str]
    query_time_ms: float
    courses: list[dict]

class SkillsUpdateRequest(BaseModel):
    skills: List[str]



class SkillGapRequest(BaseModel):
    career_goal: str
    current_skills: Optional[str] = None
    user_id: Optional[str] = None
    task_id: Optional[str] = None

class SkillGapResponse(BaseModel):
    status: str
    career_goal: str
    user_skills: List[str]
    lacking_skills: List[str]
    matched_jobs: List[dict]
    search_queries: Optional[List[dict]] = None

class CreateGanttRequest(BaseModel):
    career_goal: str
    lacking_skills: List[str]
    user_id: Optional[str] = None
    matched_jobs: Optional[List[dict]] = None
    task_id: Optional[str] = None

class GanttTaskItem(BaseModel):
    task_id: str
    phase_name: str
    skill_name: str
    course_id: str
    course_name: str
    course_url: str
    duration_weeks: int
    start_date: str
    end_date: str
    status: str = "NOT_STARTED"
    is_alternative: bool = False

class GanttChartData(BaseModel):
    tasks: List[GanttTaskItem]

class ArchitectResponse(BaseModel):
    status: str
    chart_id: Optional[str] = None
    academic_plan: str
    courses: Optional[List[dict]] = None
    alternative_courses: Optional[List[dict]] = None
    lacking_skills: Optional[List[str]] = None
    matched_jobs: Optional[List[dict]] = None
    career_goal: Optional[str] = None
    gantt_chart: Optional[dict] = None

class GetAlternativesRequest(BaseModel):
    chart_id: str
    task_id: str
    skill_name: Optional[str] = None

class SwapCourseRequest(BaseModel):
    chart_id: str
    task_id: str
    selected_course_id: str

def get_firestore_client():
    db_name = os.getenv("FIRESTORE_DATABASE")
    if db_name and db_name != "(default)":
        return firestore.Client(database=db_name)
    return firestore.Client()

LOCAL_SKILLS_DB_PATH = Path(__file__).resolve().parent / "user_skills_db.json"

def read_local_skills_db() -> dict:
    if not LOCAL_SKILLS_DB_PATH.exists():
        return {}
    try:
        with open(LOCAL_SKILLS_DB_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def write_local_skills_db(db_data: dict):
    try:
        with open(LOCAL_SKILLS_DB_PATH, "w", encoding="utf-8") as f:
            json.dump(db_data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Failed to write local skills db: {e}")

def load_jobs() -> List[dict]:
    jobs_path = Path(__file__).resolve().parent / "first_page_jobs.json"
    if not jobs_path.exists():
        logger.warning(f"Jobs file not found at {jobs_path}")
        return []
    try:
        with open(jobs_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error loading jobs file: {e}")
        return []

# Deterministic User Skill Fetching (Fixes HITL non-deterministic bug)
def fetch_user_skills_deterministic(user_id: Optional[str], current_skills: Optional[str] = None) -> List[str]:
    """Programmatically fetch user skills directly from Firestore without relying on LLM decisions."""
    user_skills = []
    if user_id:
        if USE_FIRESTORE:
            try:
                db = get_firestore_client()
                profile_doc = db.collection("profile_scanner_profiles").document(user_id).get()
                if profile_doc.exists:
                    profile_data = profile_doc.to_dict() or {}
                    user_skills = profile_data.get("skills", []) or profile_data.get("extracted_skills", [])
                    logger.info(f"Deterministic skill fetch from Firestore profile_scanner_profiles for user {user_id}: {user_skills}")
            except Exception as ex:
                logger.error(f"Failed to fetch user profile from Firestore: {ex}")

        if not user_skills:
            local_db = read_local_skills_db()
            user_skills = local_db.get(user_id, [])
            if user_skills:
                logger.info(f"Deterministic skill fetch from local DB for user {user_id}: {user_skills}")

    if not user_skills and current_skills:
        user_skills = [s.strip() for s in current_skills.split(",") if s.strip()]

    return user_skills

# Cache for embeddings instances
_embeddings_cache = {}

def get_embeddings(model_name: str = "text-embedding-004", dimensions: Optional[int] = None):
    project = os.getenv("GOOGLE_CLOUD_PROJECT") or os.getenv("PROJECT_ID")
    location = os.getenv("VERTEX_AI_LOCATION", "asia-southeast1")
    cache_key = (model_name, dimensions, project, location)
    if cache_key not in _embeddings_cache:
        kwargs = {
            "model_name": model_name,
        }
        if dimensions is not None:
            kwargs["dimensions"] = dimensions
        if project:
            kwargs["project"] = project
        if location:
            kwargs["location"] = location
        logger.info(f"Academic Architect: Initializing VertexAIEmbeddings with params {kwargs}")
        _embeddings_cache[cache_key] = VertexAIEmbeddings(**kwargs)
    return _embeddings_cache[cache_key]

# Data-Driven Job Vector Search on Firestore data_vector_embeddings (text-multilingual-embedding-002)
def perform_data_driven_job_search(career_goal: str, k: int = 30) -> tuple[List[dict], float]:
    """Retrieve top 30 candidate jobs from data_vector_embeddings and pick top 5 newest."""
    t_start = time.perf_counter()
    try:
        embeddings = get_embeddings(model_name="text-multilingual-embedding-002")
        query_vector = embeddings.embed_query(career_goal)

        db = get_firestore_client()
        col_ref = db.collection(JOBS_VECTOR_COLLECTION)

        query = col_ref.find_nearest(
            vector_field="embedding",
            query_vector=Vector(query_vector),
            distance_measure=DistanceMeasure.COSINE,
            limit=k,
            distance_result_field="vector_distance"
        )

        docs = list(query.stream())
        logger.info(f"Retrieved {len(docs)} docs from Firestore collection '{JOBS_VECTOR_COLLECTION}'")
        
        results = []
        for doc in docs:
            d = doc.to_dict() or {}
            d["document_id"] = doc.id
            results.append(d)

        if results:
            # Sort in-memory by embedding_updated_at descending to select top 5 newest
            def parse_update_time(doc_dict):
                ts = doc_dict.get("embedding_updated_at") or doc_dict.get("created_at") or ""
                return str(ts)

            results.sort(key=parse_update_time, reverse=True)
            top_5_newest = results[:5]
            elapsed_ms = (time.perf_counter() - t_start) * 1000.0
            return top_5_newest, elapsed_ms
    except Exception as ex:
        logger.warning(f"Vector search on '{JOBS_VECTOR_COLLECTION}' encountered exception/empty result: {ex}. Falling back to load_jobs().")

    # Fallback to local first_page_jobs.json if data_vector_embeddings search yields nothing
    all_jobs = load_jobs()
    goal_lower = career_goal.lower()
    goal_keywords = [w for w in re.split(r'\W+', goal_lower) if len(w) > 2]
    
    candidate_jobs = []
    for job in all_jobs:
        title = (job.get("job_title") or "").lower()
        url = (job.get("job_url") or "").lower()
        description = (job.get("Mô tả Công việc") or "").lower()
        requirements = (job.get("Yêu Cầu Công Việc") or "").lower()
        industry = (job.get("Ngành nghề") or "").lower()
        
        score = 0
        if goal_lower in title:
            score += 10
        for kw in goal_keywords:
            if kw in title:
                score += 4
            if kw in requirements:
                score += 1
        if score > 0:
            candidate_jobs.append((score, job))

    candidate_jobs.sort(key=lambda x: x[0], reverse=True)
    top_candidates = [item[1] for item in candidate_jobs[:5]]
    if not top_candidates:
        top_candidates = all_jobs[:5]

    elapsed_ms = (time.perf_counter() - t_start) * 1000.0
    return top_candidates, elapsed_ms

def perform_vector_search(search_text: str, search_mode: str, domain: Optional[str] = None, k: int = 3) -> tuple[List[dict], str]:
    """Helper to perform native Firestore vector search on Coursera catalog using find_nearest with domain pre-filtering."""
    normalized_mode = search_mode.strip().lower()
    if normalized_mode not in ("name", "description"):
        raise ValueError("search_mode must be 'name' or 'description'")
        
    if normalized_mode == "name":
        embeddings = get_embeddings(model_name="text-embedding-004", dimensions=128)
        vector_field = "name_embedding"
    else:
        embeddings = get_embeddings(model_name="text-embedding-004", dimensions=None)
        vector_field = "description_embedding"
        
    query_vector = embeddings.embed_query(search_text)
    
    db = get_firestore_client()
    col_ref = db.collection(COLLECTION_NAME)
    
    valid_domains = {
        "business", "computer-science", "information-technology", "data-science", 
        "life-sciences", "physical-science-and-engineering", "personal-development", 
        "social-sciences", "arts-and-humanities", "math-and-logic", "language-learning"
    }
    
    norm_domain = domain.strip().lower() if domain else ""
    if norm_domain not in valid_domains:
        norm_domain = "computer-science"
        
    query = col_ref.where("domainIDs", "array_contains", norm_domain).find_nearest(
        vector_field=vector_field,
        query_vector=Vector(query_vector),
        distance_measure=DistanceMeasure.COSINE,
        limit=k,
        distance_result_field="vector_distance"
    )
    
    t_firestore_start = time.perf_counter()
    docs = list(query.stream())
    t_firestore_elapsed = (time.perf_counter() - t_firestore_start) * 1000
    logger.info(f"Firestore find_nearest Coursera query took {t_firestore_elapsed:.2f} ms")
    
    results = []
    for doc in docs:
        doc_data = doc.to_dict()
        dist = doc_data.get("vector_distance")
        score = 1.0 - float(dist) if dist is not None else 0.0
        
        duration_val = doc_data.get("duration")
        if not duration_val:
            desc_lower = (doc_data.get("description") or "").lower()
            name_lower = (doc_data.get("name") or "").lower()
            text_to_search = name_lower + " " + desc_lower
            
            match = re.search(r'(\d+)\s*(hour|hr|giờ|hrs)', text_to_search)
            if match:
                duration_val = f"{match.group(1)} giờ"
            else:
                certs_list = doc_data.get("certificates") or []
                if any("specialization" in str(c).lower() for c in certs_list):
                    duration_val = "3 tháng"
                elif any("professional" in str(c).lower() for c in certs_list):
                    duration_val = "6 tháng"
                else:
                    duration_val = "15 giờ"

        course = {
            "course_id": doc_data.get("course_id") or doc.id,
            "name": doc_data.get("name") or "",
            "description": doc_data.get("description") or "",
            "slug": doc_data.get("slug") or "",
            "course_type": doc_data.get("course_type") or "",
            "certificates": doc_data.get("certificates") or [],
            "domain_types": doc_data.get("domain_types") or [],
            "partners": doc_data.get("partners") or [],
            "photo_url": doc_data.get("photo_url") or "",
            "score": score,
            "duration": duration_val,
            "workload": doc_data.get("workload") or ""
        }
        
        slug = course.get("slug")
        course["url"] = f"https://www.coursera.org/learn/{slug}" if slug else ""
        results.append(course)
        
    return results, norm_domain

# --- SSE Progress Stream Endpoint ---
@app.get("/stream-progress/{task_id}")
async def stream_progress(task_id: str, request: Request):
    """Server-Sent Events endpoint to stream execution progress logs with state timings."""
    if task_id not in progress_queues:
        progress_queues[task_id] = asyncio.Queue()

    async def event_generator():
        q = progress_queues[task_id]
        while True:
            if await request.is_disconnected():
                logger.info(f"SSE client disconnected for task {task_id}")
                break
            try:
                event_data = await asyncio.wait_for(q.get(), timeout=1.0)
                yield {
                    "event": "progress",
                    "data": json.dumps(event_data, ensure_ascii=False)
                }
            except asyncio.TimeoutError:
                # Send heartbeat
                yield {
                    "event": "heartbeat",
                    "data": json.dumps({"timestamp": datetime.datetime.utcnow().isoformat() + "Z"})
                }

    return EventSourceResponse(event_generator())

# --- Endpoint 1: USE CASE 1 - /skill-gap ---
@app.post("/skill-gap", response_model=SkillGapResponse)
async def analyze_skill_gap(request: SkillGapRequest):
    """Use Case 1: Data-driven skill gap analysis matching top 5 newest jobs from 30 vector-matched postings."""
    t_total_start = time.perf_counter()
    task_id = request.task_id or f"task_{uuid.uuid4().hex[:8]}"

    # Step 1: Deterministic Skill Fetching
    t_step_start = time.perf_counter()
    user_skills = fetch_user_skills_deterministic(request.user_id, request.current_skills)
    step_duration = (time.perf_counter() - t_step_start) * 1000.0
    await emit_progress(task_id, "FETCH_USER_SKILLS", "COMPLETED", step_duration, f"Programmatically retrieved user skills: {user_skills}")

    # Step 2: Data-Driven Job Vector Search (text-multilingual-embedding-002)
    await emit_progress(task_id, "VECTOR_SEARCH_JOBS", "STARTED", 0.0, f"Performing vector search on 'data_vector_embeddings' for career goal '{request.career_goal}'...")
    top_5_jobs, vector_duration = await asyncio.to_thread(perform_data_driven_job_search, request.career_goal, 30)
    await emit_progress(task_id, "VECTOR_SEARCH_JOBS", "COMPLETED", vector_duration, f"Retrieved 30 jobs and selected top 5 newest postings in {vector_duration:.1f}ms.")

    # Step 3: LLM Skill Gap Extraction from real job texts
    t_llm_start = time.perf_counter()
    await emit_progress(task_id, "SKILL_GAP_ANALYSIS", "STARTED", 0.0, "Extracting requirements from real job postings and analyzing skill gaps...")

    formatted_jobs_context = ""
    matched_jobs_list = []
    for idx, job in enumerate(top_5_jobs):
        title = job.get("job_title") or job.get("title") or "Tuyển dụng"
        company = job.get("company") or job.get("Company") or "Công ty tuyển dụng"
        salary = job.get("salary") or job.get("Lương") or "Thỏa thuận"
        url = job.get("job_url") or job.get("url") or ""
        emb_text = job.get("embedding_text") or job.get("Mô tả Công việc") or job.get("Yêu Cầu Công Việc") or ""
        
        matched_jobs_list.append({
            "job_id": job.get("job_id") or f"job_{idx}",
            "job_title": title,
            "company": company,
            "salary": salary,
            "url": url
        })
        formatted_jobs_context += f"Job Index {idx}:\nTitle: {title}\nCompany: {company}\nSalary: {salary}\nDetails: {emb_text[:500]}\nURL: {url}\n\n"

    prompt_skill_gap = f"""
You are the AI Academic Architect.
Target Career Goal: "{request.career_goal}"
User's Current Skills: {json.dumps(user_skills)}

Here are 5 real, recent job market postings matching the career goal:
{formatted_jobs_context}

Step 1: Extract the specific technical and domain requirements from these real job postings.
Step 2: Compare these target requirements against the user's current skills. Identify the specific skills the user is lacking.
Step 3: For each lacking skill, formulate a single search query to look up courses in Coursera.
Specify a domain from one of: "business", "computer-science", "information-technology", "data-science", "life-sciences", "physical-science-and-engineering", "personal-development", "social-sciences", "arts-and-humanities", "math-and-logic", "language-learning"

Return valid JSON with:
- "matched_jobs": list of objects with "job_id", "job_title", "company", "url"
- "lacking_skills": list of strings of missing skills
- "search_queries": list of objects containing "subject", "search_text", "search_mode" ("name" or "description"), "domain"

Do not wrap in markdown outside JSON.
"""
    if llm:
        res = await llm.ainvoke(prompt_skill_gap)
        raw_content = res.content
        if isinstance(raw_content, list):
            raw_content = "".join([p.get("text", "") if isinstance(p, dict) else str(p) for p in raw_content])
        raw_content = raw_content.strip()
        if raw_content.startswith("```json"):
            raw_content = raw_content[7:]
        if raw_content.endswith("```"):
            raw_content = raw_content[:-3]
        raw_content = raw_content.strip()

        try:
            analysis_data = json.loads(raw_content)
        except Exception:
            logger.warning("Failed to parse LLM skill gap JSON, using default structure.")
            analysis_data = {
                "matched_jobs": matched_jobs_list,
                "lacking_skills": ["Chuyên môn bổ trợ"],
                "search_queries": [{"subject": "Skill", "search_text": request.career_goal, "search_mode": "description", "domain": "computer-science"}]
            }
    else:
        analysis_data = {
            "matched_jobs": matched_jobs_list,
            "lacking_skills": ["Kỹ năng chuyên môn"],
            "search_queries": [{"subject": "Skill", "search_text": request.career_goal, "search_mode": "description", "domain": "computer-science"}]
        }

    llm_duration = (time.perf_counter() - t_llm_start) * 1000.0
    lacking_skills = analysis_data.get("lacking_skills", [])
    await emit_progress(task_id, "SKILL_GAP_ANALYSIS", "COMPLETED", llm_duration, f"Skill gap analysis completed. Lacking skills identified: {lacking_skills}")

    return SkillGapResponse(
        status="success",
        career_goal=request.career_goal,
        user_skills=user_skills,
        lacking_skills=lacking_skills,
        matched_jobs=analysis_data.get("matched_jobs", matched_jobs_list),
        search_queries=analysis_data.get("search_queries", [])
    )

# --- Endpoint 2: USE CASE 2 - /create-gantt ---
@app.post("/create-gantt", response_model=ArchitectResponse)
async def create_gantt_roadmap(request: CreateGanttRequest):
    """Use Case 2: Query Coursera catalog, build structured Gantt timeline & narrative text, and store in Firestore."""
    task_id = request.task_id or f"task_{uuid.uuid4().hex[:8]}"
    t_start = time.perf_counter()

    await emit_progress(task_id, "COURSE_SEARCH", "STARTED", 0.0, "Searching Coursera vector catalog for missing skills...")
    
    selected_courses = []
    alternative_courses = []
    seen_course_ids = set()
    gantt_tasks = []

    # Map lacking skills to Coursera vector searches
    current_start_date = datetime.date.today() + datetime.timedelta(days=7) # Start next week

    for idx, skill in enumerate(request.lacking_skills):
        t_sub = time.perf_counter()
        results, domain_used = await asyncio.to_thread(
            perform_vector_search, skill, "description", "computer-science", 4
        )
        dur_sub = (time.perf_counter() - t_sub) * 1000.0
        
        if results:
            best = results[0]
            if best["course_id"] not in seen_course_ids:
                seen_course_ids.add(best["course_id"])
                selected_courses.append(best)

            for alt in results[1:]:
                if alt["course_id"] not in seen_course_ids:
                    seen_course_ids.add(alt["course_id"])
                    alternative_courses.append(alt)

            # Build Gantt task
            duration_weeks = 4 # Default 4 weeks per course module
            if "giờ" in best.get("duration", ""):
                try:
                    hrs = int(re.search(r'\d+', best["duration"]).group(0))
                    duration_weeks = max(1, round(hrs / 10))
                except Exception:
                    pass
            elif "tháng" in best.get("duration", ""):
                try:
                    m = int(re.search(r'\d+', best["duration"]).group(0))
                    duration_weeks = m * 4
                except Exception:
                    pass

            end_date = current_start_date + datetime.timedelta(weeks=duration_weeks)

            task_item = {
                "task_id": f"task_{idx+1}",
                "phase_name": f"Phase {idx+1}: {skill}",
                "skill_name": skill,
                "course_id": best["course_id"],
                "course_name": best["name"],
                "course_url": best["url"],
                "duration_weeks": duration_weeks,
                "start_date": current_start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "status": "NOT_STARTED",
                "is_alternative": False
            }
            gantt_tasks.append(task_item)
            current_start_date = end_date + datetime.timedelta(days=1)

    # Generate academic plan narrative text
    t_plan_start = time.perf_counter()
    await emit_progress(task_id, "GENERATING_NARRATIVE", "STARTED", 0.0, "Generating academic plan narrative text...")

    narrative_prompt = f"""
Hãy tạo một bài thuyết minh lộ trình học tập (Markdown format) cho mục tiêu nghề nghiệp: "{request.career_goal}".
Danh sách các kỹ năng cần học và khóa học tương ứng:
{json.dumps(gantt_tasks, ensure_ascii=False, indent=2)}

Bài viết cần chia theo các Phase rõ ràng (Phase 1, Phase 2...), mô tả ngắn gọn vai trò của từng kỹ năng và lý do chọn khóa học.
Format bài viết bằng GitHub Markdown đẹp mắt.
"""
    if llm:
        res_narrative = await llm.ainvoke(narrative_prompt)
        academic_plan_narrative = res_narrative.content.strip()
    else:
        academic_plan_narrative = f"### Lộ Trình Học Tập Cho {request.career_goal}\n\nCác giai đoạn học tập đã được thiết lập dựa trên kỹ năng còn thiếu."

    narrative_duration = (time.perf_counter() - t_plan_start) * 1000.0

    # Build chart_id & save document in Firestore
    chart_id = f"gantt_{uuid.uuid4().hex[:10]}"
    chart_payload = {
        "chart_id": chart_id,
        "user_id": request.user_id,
        "career_goal": request.career_goal,
        "created_at": datetime.datetime.utcnow().isoformat() + "Z",
        "academic_plan_narrative": academic_plan_narrative,
        "matched_jobs": request.matched_jobs or [],
        "tasks": gantt_tasks
    }

    if USE_FIRESTORE:
        try:
            db = get_firestore_client()
            db.collection(GANTT_COLLECTION_NAME).document(chart_id).set(chart_payload)
            logger.info(f"Saved Gantt chart to Firestore collection '{GANTT_COLLECTION_NAME}' with ID: {chart_id}")
        except Exception as ex:
            logger.error(f"Failed to save Gantt chart in Firestore: {ex}")
    else:
        try:
            db_local = read_gantt_db()
            db_local[chart_id] = chart_payload
            write_gantt_db(db_local)
            logger.info(f"Saved Gantt chart to local JSON with ID: {chart_id}")
        except Exception as ex:
            logger.error(f"Failed to save Gantt chart to local JSON: {ex}")

    total_duration = (time.perf_counter() - t_start) * 1000.0
    await emit_progress(task_id, "GANTT_ROADMAP_COMPLETE", "COMPLETED", total_duration, f"Gantt roadmap generated and saved successfully. Chart ID: {chart_id}")

    return ArchitectResponse(
        status="success",
        chart_id=chart_id,
        academic_plan=academic_plan_narrative,
        courses=selected_courses,
        alternative_courses=alternative_courses,
        lacking_skills=request.lacking_skills,
        matched_jobs=request.matched_jobs or [],
        career_goal=request.career_goal,
        gantt_chart={"tasks": gantt_tasks}
    )

# --- Endpoint 3: COURSE SWAP USE CASE - Step 1: /get-alternatives ---
@app.post("/chart/{chart_id}/get-alternatives")
async def get_course_alternatives(chart_id: str, request: GetAlternativesRequest):
    """Course Swap Step 1: Search top 5 alternative courses in Coursera vector database for a specified task."""
    db = get_firestore_client() if USE_FIRESTORE else None
    doc_ref = db.collection(GANTT_COLLECTION_NAME).document(chart_id).get() if USE_FIRESTORE else None
    
    skill_query = request.skill_name or "Chuyên môn"
    current_course_id = ""

    if doc_ref and doc_ref.exists:
        chart_data = doc_ref.to_dict() or {}
        tasks = chart_data.get("tasks", [])
        for t in tasks:
            if t.get("task_id") == request.task_id:
                skill_query = t.get("skill_name") or skill_query
                current_course_id = t.get("course_id") or ""
                break
    elif not USE_FIRESTORE:
        db_local = read_gantt_db()
        if chart_id in db_local:
            chart_data = db_local[chart_id]
            tasks = chart_data.get("tasks", [])
            for t in tasks:
                if t.get("task_id") == request.task_id:
                    skill_query = t.get("skill_name") or skill_query
                    current_course_id = t.get("course_id") or ""
                    break

    # Vector search top 6 courses to exclude the current course
    results, _ = await asyncio.to_thread(perform_vector_search, skill_query, "description", "computer-science", 6)
    filtered_alts = [c for c in results if c["course_id"] != current_course_id][:5]

    return {
        "status": "success",
        "chart_id": chart_id,
        "task_id": request.task_id,
        "skill_name": skill_query,
        "alternatives": filtered_alts
    }

# --- Endpoint 4: COURSE SWAP USE CASE - Step 2: /swap-course ---
@app.post("/chart/{chart_id}/swap-course")
async def swap_course(chart_id: str, request: SwapCourseRequest):
    """Course Swap Step 2: Replace specified task course with selected alternative ID and recalculate Gantt timeline."""
    if USE_FIRESTORE:
        db = get_firestore_client()
        doc_ref = db.collection(GANTT_COLLECTION_NAME).document(chart_id)
        doc = doc_ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail=f"Gantt chart '{chart_id}' not found.")
        chart_data = doc.to_dict() or {}
    else:
        db_local = read_gantt_db()
        if chart_id not in db_local:
            raise HTTPException(status_code=404, detail=f"Gantt chart '{chart_id}' not found.")
        chart_data = db_local[chart_id]

    tasks = chart_data.get("tasks", [])

    selected_name = "Alternative Course"
    selected_url = ""

    if USE_FIRESTORE:
        # Fetch selected course details from Coursera catalog
        col_ref = db.collection(COLLECTION_NAME)
        course_doc = col_ref.document(request.selected_course_id).get()
        if course_doc.exists:
            c_data = course_doc.to_dict() or {}
            selected_name = c_data.get("name") or selected_name
            slug = c_data.get("slug")
            selected_url = f"https://www.coursera.org/learn/{slug}" if slug else ""
    else:
        # If no firestore, do a quick vector search to get the details
        results, _ = await asyncio.to_thread(perform_vector_search, request.selected_course_id, "name", "computer-science", 10)
        for c in results:
            if c["course_id"] == request.selected_course_id:
                selected_name = c.get("name") or selected_name
                selected_url = c.get("url") or ""
                break

    # Find and update task
    task_found = False
    for t in tasks:
        if t.get("task_id") == request.task_id:
            t["course_id"] = request.selected_course_id
            t["course_name"] = selected_name
            t["course_url"] = selected_url
            t["is_alternative"] = True
            task_found = True
            break

    if not task_found:
        raise HTTPException(status_code=404, detail=f"Task '{request.task_id}' not found in chart.")

    # Recalculate start and end dates for all tasks
    start_date = datetime.date.today() + datetime.timedelta(days=7)
    for t in tasks:
        dur = t.get("duration_weeks", 4)
        end_date = start_date + datetime.timedelta(weeks=dur)
        t["start_date"] = start_date.isoformat()
        t["end_date"] = end_date.isoformat()
        start_date = end_date + datetime.timedelta(days=1)

    chart_data["tasks"] = tasks
    chart_data["updated_at"] = datetime.datetime.utcnow().isoformat() + "Z"
    doc_ref.set(chart_data, merge=True)

    return {
        "status": "success",
        "chart_id": chart_id,
        "updated_task_id": request.task_id,
        "gantt_chart": {"tasks": tasks}
    }

GANTT_DB_PATH = os.path.join(os.path.dirname(__file__), "gantt_db.json")

def read_gantt_db() -> dict:
    if not os.path.exists(GANTT_DB_PATH):
        return {}
    try:
        with open(GANTT_DB_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def write_gantt_db(db: dict):
    with open(GANTT_DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, indent=2, ensure_ascii=False)

# --- Endpoint 5: Backend Excel Generation (/chart/{chart_id}/excel) ---
@app.get("/chart/{chart_id}/excel")
async def export_gantt_excel(chart_id: str):
    """Generate and stream an openpyxl Excel (.xlsx) workbook for the specified Gantt roadmap."""
    if USE_FIRESTORE:
        db = get_firestore_client()
        doc = db.collection(GANTT_COLLECTION_NAME).document(chart_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail=f"Gantt chart '{chart_id}' not found.")
        chart_data = doc.to_dict() or {}
    else:
        db_local = read_gantt_db()
        if chart_id not in db_local:
            raise HTTPException(status_code=404, detail=f"Gantt chart '{chart_id}' not found.")
        chart_data = db_local[chart_id]
    tasks = chart_data.get("tasks", [])
    career_goal = chart_data.get("career_goal", "Lộ Trình Học Tập")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Roadmap"

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(name="Arial", size=14, bold=True, color="1F4E78")
    cell_font = Font(name="Arial", size=10)
    gantt_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = f"LỘ TRÌNH HỌC TẬP: {career_goal.upper()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Parse dates to calculate project timeline
    for task in tasks:
        sd = task.get("start_date", "")
        ed = task.get("end_date", "")
        try:
            task["_parsed_start"] = datetime.datetime.fromisoformat(sd.replace("Z", "+00:00")).date() if sd else datetime.date.today()
            task["_parsed_end"] = datetime.datetime.fromisoformat(ed.replace("Z", "+00:00")).date() if ed else datetime.date.today() + datetime.timedelta(days=28)
        except Exception:
            task["_parsed_start"] = datetime.date.today()
            task["_parsed_end"] = datetime.date.today() + datetime.timedelta(days=28)

    if tasks:
        earliest_start = min(t["_parsed_start"] for t in tasks)
        latest_end = max(t["_parsed_end"] for t in tasks)
    else:
        earliest_start = datetime.date.today()
        latest_end = datetime.date.today() + datetime.timedelta(days=30)
    
    # Generate 3-day interval buckets
    timeline_dates = []
    curr = earliest_start
    while curr <= latest_end:
        timeline_dates.append(curr)
        curr += datetime.timedelta(days=3)
    
    # Ensure the last bucket covers the latest_end if it doesn't align perfectly
    if not timeline_dates or timeline_dates[-1] < latest_end:
        timeline_dates.append(curr)

    total_buckets = len(timeline_dates)

    headers = ["Task ID", "Giai Đoạn (Phase)", "Kỹ Năng Mục Tiêu", "Tên Khóa Học", "Thời Gian (Tuần)", "Ngày Bắt Đầu", "Ngày Kết Thúc", "Trạng Thái", "Đường Dẫn Khóa Học"]
    base_header_len = len(headers)
    
    for d in timeline_dates:
        headers.append(d.strftime("%b %d"))

    ws.append([])
    ws.append(headers)

    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, task in enumerate(tasks, start=4):
        row_data = [
            task.get("task_id", ""),
            task.get("phase_name", ""),
            task.get("skill_name", ""),
            task.get("course_name", ""),
            task.get("duration_weeks", 4),
            task.get("start_date", ""),
            task.get("end_date", ""),
            task.get("status", "NOT_STARTED"),
            task.get("course_url", "")
        ]
        
        # Add empty cells for the timeline section
        row_data.extend([""] * total_buckets)
        ws.append(row_data)

        t_start = task["_parsed_start"]
        t_end = task["_parsed_end"]

        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = cell_font
            cell.border = thin_border
            
            if col_num <= base_header_len:
                if col_num in (1, 5, 6, 7, 8):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                bucket_idx = col_num - base_header_len - 1
                b_start = timeline_dates[bucket_idx]
                b_end = b_start + datetime.timedelta(days=2) # 3 days inclusive

                # Overlap condition: task starts on or before bucket ends AND task ends on or after bucket starts
                if t_start <= b_end and t_end >= b_start:
                    cell.fill = gantt_fill

    for col in ws.columns:
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        if col_idx <= base_header_len:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        else:
            ws.column_dimensions[col_letter].width = 8

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"gantt_roadmap_{chart_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@app.get("/skills/{user_id}")
async def get_user_skills(user_id: str):
    user_skills = fetch_user_skills_deterministic(user_id)
    return {"skills": user_skills}

@app.post("/skills/{user_id}")
async def update_user_skills(user_id: str, request: SkillsUpdateRequest):
    updated_firestore = False
    if USE_FIRESTORE:
        try:
            db = get_firestore_client()
            doc_ref = db.collection("profile_scanner_profiles").document(user_id)
            updates = {
                "skills": request.skills,
                "extracted_skills": request.skills,
                "updated_at": datetime.datetime.utcnow().isoformat() + "Z"
            }
            doc_ref.set(updates, merge=True)
            logger.info(f"Updated user skills in Firestore profile_scanner_profiles for user {user_id}")
            updated_firestore = True
        except Exception as e:
            logger.error(f"Failed to update user skills in Firestore: {e}")
            raise HTTPException(status_code=500, detail=str(e))
            
    try:
        local_db = read_local_skills_db()
        local_db[user_id] = request.skills
        write_local_skills_db(local_db)
        logger.info(f"Updated user skills in local DB for user {user_id}: {request.skills}")
    except Exception as e:
        logger.error(f"Failed to update user skills in local DB: {e}")
        if not updated_firestore:
            raise HTTPException(status_code=500, detail=str(e))
            
    return {"status": "success", "skills": request.skills}

@app.get("/health")
async def health_check():
    return {
        "status": "ok", 
        "service": "academic_architect"
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8080)
